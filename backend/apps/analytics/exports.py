"""
File:    backend/apps/analytics/exports.py
Purpose: Build report payloads + render PDF and XLSX bytes.
Owner:   Navanish

Design:
  - One `assemble_*` function per scope (student / class / school) — returns a
    plain dict so the same data can drive PDF, XLSX, JSON, or future channels.
  - Two stateless renderers — `render_pdf(report)` and `render_xlsx(report)`.
  - Views are thin: they assemble the payload, pick the renderer, and stream
    the bytes as an HttpResponse. No business logic in views.

The renderers depend on `reportlab` and `openpyxl` (added to requirements in
Phase 1.5). If either import fails the export endpoint will surface a clean
500 with `pip install` guidance — we deliberately do NOT silently degrade.
"""

from __future__ import annotations

import datetime as _dt
import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from django.db.models import Avg, Sum

from apps.academics.models import Class
from apps.accounts.models import User
from apps.schools.models import School

from .models import ClassWeeklyStats, StudentDailyStats


# ── Data assembly ────────────────────────────────────────────────────────────


@dataclass
class DateRange:
    start: _dt.date
    end: _dt.date

    @classmethod
    def last_n_days(cls, n: int) -> "DateRange":
        end = _dt.date.today()
        return cls(start=end - _dt.timedelta(days=n - 1), end=end)


def assemble_student_report(
    *, school_id, student: User, date_range: DateRange
) -> dict[str, Any]:
    """One student's day-by-day stats over the given window."""
    rows = list(
        StudentDailyStats.objects
        .filter(
            school_id=school_id,
            student=student,
            date__gte=date_range.start,
            date__lte=date_range.end,
        )
        .order_by("date")
        .values("date", "quizzes_taken", "avg_score", "time_spent_seconds")
    )
    agg = StudentDailyStats.objects.filter(
        school_id=school_id, student=student,
        date__gte=date_range.start, date__lte=date_range.end,
    ).aggregate(avg=Avg("avg_score"), time=Sum("time_spent_seconds"))

    total_quizzes = sum(int(r["quizzes_taken"]) for r in rows)
    avg_score = _to_float(agg["avg"]) or 0.0
    total_time_min = (agg["time"] or 0) / 60.0

    return {
        "kind": "student",
        "title": f"Student Progress — {student.get_full_name() or student.username}",
        "subtitle": f"{student.email} · {date_range.start} → {date_range.end}",
        "summary": [
            ("Total quizzes taken",       f"{total_quizzes}"),
            ("Average score",             f"{avg_score:.2f}%"),
            ("Time spent (minutes)",      f"{total_time_min:.1f}"),
            ("Days with activity",        f"{len(rows)}"),
        ],
        "table": {
            "headers": ["Date", "Quizzes", "Avg Score (%)", "Time (min)"],
            "rows": [
                [
                    r["date"].isoformat(),
                    r["quizzes_taken"],
                    f"{_to_float(r['avg_score']):.2f}",
                    f"{(r['time_spent_seconds'] or 0) / 60:.1f}",
                ]
                for r in rows
            ],
        },
    }


def assemble_class_report(
    *, school_id, klass: Class, date_range: DateRange
) -> dict[str, Any]:
    """Class-level weekly aggregates over the window."""
    rows = list(
        ClassWeeklyStats.objects
        .filter(
            school_id=school_id,
            klass=klass,
            week_start_date__gte=date_range.start,
            week_start_date__lte=date_range.end,
        )
        .order_by("week_start_date")
        .values("week_start_date", "avg_score", "at_risk_count")
    )
    agg = ClassWeeklyStats.objects.filter(
        school_id=school_id, klass=klass,
        week_start_date__gte=date_range.start, week_start_date__lte=date_range.end,
    ).aggregate(avg=Avg("avg_score"))

    return {
        "kind": "class",
        "title": f"Class Report — Grade {klass.grade}-{klass.section}",
        "subtitle": f"Academic year {klass.academic_year.name} · {date_range.start} → {date_range.end}",
        "summary": [
            ("Average score (window)",    f"{(_to_float(agg['avg']) or 0):.2f}%"),
            ("Weeks tracked",             f"{len(rows)}"),
            ("At-risk students (peak)",   f"{max((r['at_risk_count'] for r in rows), default=0)}"),
        ],
        "table": {
            "headers": ["Week Starting", "Avg Score (%)", "At-Risk Students"],
            "rows": [
                [
                    r["week_start_date"].isoformat(),
                    f"{_to_float(r['avg_score']):.2f}",
                    r["at_risk_count"],
                ]
                for r in rows
            ],
        },
    }


def assemble_school_report(
    *, school: School, date_range: DateRange
) -> dict[str, Any]:
    """School-wide rollup: per-class averages + at-risk count over the window."""
    per_class = list(
        ClassWeeklyStats.objects
        .filter(
            school=school,
            week_start_date__gte=date_range.start,
            week_start_date__lte=date_range.end,
        )
        .values("klass__grade", "klass__section")
        .annotate(avg=Avg("avg_score"), risk=Sum("at_risk_count"))
        .order_by("klass__grade", "klass__section")
    )
    overall_avg = (
        ClassWeeklyStats.objects.filter(
            school=school,
            week_start_date__gte=date_range.start,
            week_start_date__lte=date_range.end,
        ).aggregate(avg=Avg("avg_score"))["avg"]
    )

    return {
        "kind": "school",
        "title": f"School Report — {school.name}",
        "subtitle": f"{school.board} · {school.city}, {school.state} · {date_range.start} → {date_range.end}",
        "summary": [
            ("Overall average",           f"{(_to_float(overall_avg) or 0):.2f}%"),
            ("Classes covered",           f"{len(per_class)}"),
            ("Total at-risk events",      f"{sum(int(r['risk'] or 0) for r in per_class)}"),
        ],
        "table": {
            "headers": ["Class", "Avg Score (%)", "At-Risk Events"],
            "rows": [
                [
                    f"Grade {r['klass__grade']}-{r['klass__section']}",
                    f"{_to_float(r['avg']) or 0:.2f}",
                    int(r["risk"] or 0),
                ]
                for r in per_class
            ],
        },
    }


# ── Renderers ────────────────────────────────────────────────────────────────


def render_pdf(report: dict[str, Any]) -> bytes:
    """Render the assembled report as a single-page A4 PDF using reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError as exc:  # pragma: no cover — fail loud on missing dep
        raise RuntimeError(
            "reportlab is required for PDF export. Run: pip install reportlab"
        ) from exc

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=report["title"],
    )
    styles = getSampleStyleSheet()
    story: list = []

    story.append(Paragraph(report["title"], styles["Heading1"]))
    story.append(Paragraph(report["subtitle"], styles["Italic"]))
    story.append(Spacer(1, 0.6 * cm))

    # Summary KPIs
    summary_table = Table(
        report["summary"], colWidths=[6 * cm, 8 * cm], hAlign="LEFT",
    )
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#475569")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.8 * cm))

    # Detail table
    tdata = [report["table"]["headers"], *report["table"]["rows"]]
    if len(tdata) == 1:
        story.append(Paragraph("<i>No data available for the selected window.</i>", styles["BodyText"]))
    else:
        detail_table = Table(tdata, repeatRows=1, hAlign="LEFT")
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(detail_table)

    doc.build(story)
    return buf.getvalue()


def render_xlsx(report: dict[str, Any]) -> bytes:
    """Render the report as a 2-sheet .xlsx file (Summary + Detail)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for Excel export. Run: pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"

    # Header rows
    ws1["A1"] = report["title"]
    ws1["A1"].font = Font(size=14, bold=True)
    ws1["A2"] = report["subtitle"]
    ws1["A2"].font = Font(italic=True, color="475569")

    ws1.append([])  # blank row
    for label, value in report["summary"]:
        ws1.append([label, value])

    for cell in ws1["A"]:
        cell.font = Font(bold=cell.font.bold or False)
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 24

    # Detail sheet
    ws2 = wb.create_sheet("Detail")
    headers = report["table"]["headers"]
    ws2.append(headers)
    header_fill = PatternFill("solid", fgColor="0F766E")
    for col_idx in range(1, len(headers) + 1):
        c = ws2.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")

    for row in report["table"]["rows"]:
        ws2.append(row)

    for col_idx in range(1, len(headers) + 1):
        col_letter = ws2.cell(row=1, column=col_idx).column_letter
        ws2.column_dimensions[col_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ──────────────────────────────────────────────────────────────────


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return float(value)
